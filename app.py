import os, re, hashlib, csv, io
from flask import Flask, render_template, jsonify, request, send_file
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

try:
    import psycopg2, psycopg2.extras
    PG_OK = True
except ImportError:
    PG_OK = False

try:
    import cloudinary, cloudinary.uploader
    CLOUD_OK = True
except ImportError:
    CLOUD_OK = False

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

DATABASE_URL   = os.getenv("DATABASE_URL", "")
CLOUDINARY_URL = os.getenv("CLOUDINARY_URL", "")

# Render gives postgres:// but psycopg2 needs postgresql://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS candidates (
        id                 SERIAL PRIMARY KEY,
        upload_batch       TEXT,
        recruiter_name     TEXT,
        recruiter_email    TEXT,
        role_name          TEXT,
        candidate_name     TEXT,
        email_addr         TEXT,
        phone              TEXT,
        current_company    TEXT,
        current_role       TEXT,
        experience_years   TEXT,
        key_skills         TEXT,
        notice_period      TEXT,
        current_salary     TEXT,
        expected_salary    TEXT,
        current_location   TEXT,
        preferred_location TEXT,
        remarks            TEXT,
        cv_filename        TEXT,
        cv_url             TEXT,
        cv_public_id       TEXT,
        is_duplicate       INTEGER DEFAULT 0,
        duplicate_of       INTEGER,
        missing_info       TEXT,
        created_at         TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS team_members (
        id        SERIAL PRIMARY KEY,
        name      TEXT,
        email     TEXT UNIQUE,
        is_fixed  INTEGER DEFAULT 0,
        added_at  TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS alerts (
        id              SERIAL PRIMARY KEY,
        alert_type      TEXT,
        message         TEXT,
        candidate_id    INTEGER,
        recruiter_email TEXT,
        is_read         INTEGER DEFAULT 0,
        created_at      TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS upload_log (
        id               SERIAL PRIMARY KEY,
        batch_id         TEXT,
        recruiter_name   TEXT,
        recruiter_email  TEXT,
        filename         TEXT,
        candidates_added INTEGER DEFAULT 0,
        duplicates_found INTEGER DEFAULT 0,
        missing_count    INTEGER DEFAULT 0,
        uploaded_at      TIMESTAMP DEFAULT NOW()
    );
    """)
    conn.commit(); cur.close(); conn.close()

# ── Column map ────────────────────────────────────────────────────────────────
COL_MAP = {
    "candidate name":"candidate_name","name":"candidate_name","candidate":"candidate_name",
    "email":"email_addr","email id":"email_addr","email address":"email_addr",
    "phone":"phone","mobile":"phone","contact":"phone","contact no":"phone","contact number":"phone",
    "current company":"current_company","company":"current_company","employer":"current_company",
    "organisation":"current_company","organization":"current_company",
    "current role":"current_role","designation":"current_role","title":"current_role",
    "position":"current_role","current designation":"current_role",
    "experience":"experience_years","exp":"experience_years","total exp":"experience_years",
    "years of exp":"experience_years","experience (years)":"experience_years",
    "yrs":"experience_years","total experience":"experience_years",
    "skills":"key_skills","key skills":"key_skills","skill set":"key_skills",
    "tech skills":"key_skills","technical skills":"key_skills",
    "notice":"notice_period","notice period":"notice_period","np":"notice_period",
    "current salary":"current_salary","current ctc":"current_salary","ctc":"current_salary","salary":"current_salary",
    "expected salary":"expected_salary","expected ctc":"expected_salary","ectc":"expected_salary","exp salary":"expected_salary",
    "location":"current_location","current location":"current_location","city":"current_location",
    "preferred location":"preferred_location","pref location":"preferred_location","preferred city":"preferred_location",
    "remarks":"remarks","notes":"remarks","comments":"remarks","feedback":"remarks",
    "role":"role_name","job role":"role_name","applied for":"role_name",
    "position applied":"role_name","job title":"role_name",
}
ALL_FIELDS = ["candidate_name","email_addr","phone","current_company","current_role",
              "experience_years","key_skills","notice_period","current_salary",
              "expected_salary","current_location","preferred_location","remarks","role_name"]

def norm_key(k): return COL_MAP.get(str(k).strip().lower(), None)
def empty_row(): return {f: "" for f in ALL_FIELDS}

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_xlsx(file_bytes, role_override=""):
    if not XLSX_OK: return [], "openpyxl not installed"
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        if not data: continue
        hdr_idx = 0
        for i, row in enumerate(data):
            if len([c for c in row if c is not None and str(c).strip()]) >= 3:
                hdr_idx = i; break
        headers = [norm_key(str(c).strip() if c else "") for c in data[hdr_idx]]
        for dr in data[hdr_idx+1:]:
            if not any(c is not None and str(c).strip() for c in dr): continue
            row = empty_row()
            for i, cell in enumerate(dr):
                if i >= len(headers): break
                key = headers[i]
                val = str(cell).strip() if cell is not None else ""
                if key and key in ALL_FIELDS: row[key] = val
            if role_override and not row["role_name"]: row["role_name"] = role_override
            rows.append(row)
    return rows, None

def parse_csv(file_bytes, role_override=""):
    rows   = []
    text   = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    for dr in reader:
        row = empty_row()
        for k, v in dr.items():
            key = norm_key(k)
            if key and key in ALL_FIELDS: row[key] = (v or "").strip()
        if role_override and not row["role_name"]: row["role_name"] = role_override
        rows.append(row)
    return rows, None

# ── CV upload (Cloudinary) ────────────────────────────────────────────────────
def upload_cv(file_storage, batch_id):
    fname = secure_filename(file_storage.filename)
    ext   = os.path.splitext(fname)[1].lower()
    if ext not in (".pdf",".doc",".docx"): return fname, None, None
    if not CLOUD_OK or not CLOUDINARY_URL:  return fname, None, None
    cloudinary.config(cloudinary_url=CLOUDINARY_URL)
    result = cloudinary.uploader.upload(
        file_storage, folder="hrguru_cvs",
        public_id=f"{batch_id}_{fname}",
        resource_type="raw", use_filename=True, unique_filename=False)
    return fname, result.get("secure_url"), result.get("public_id")

def match_cv(name, cv_files):
    if not cv_files: return None, None, None, False
    parts = [p.lower() for p in (name or "").split() if len(p) > 2]
    best_score, best = 0, (None, None, None)
    for orig, url, pub_id in cv_files:
        score = sum(1 for p in parts if p in orig.lower())
        if score > best_score: best_score, best = score, (orig, url, pub_id)
    if best_score > 0: return *best, True
    return None, None, None, False

def unmatched_cvs(cv_files, matched_ids):
    return [(o,u,p) for o,u,p in cv_files if p not in matched_ids]

# ── Dedup ─────────────────────────────────────────────────────────────────────
def norm_phone(p): return re.sub(r'[^\d]','',p or "")[-10:]

def check_dup(conn, row):
    cur   = conn.cursor()
    phone = norm_phone(row.get("phone",""))
    email = (row.get("email_addr","") or "").strip().lower()
    if phone and len(phone) >= 8:
        cur.execute("""SELECT id FROM candidates
            WHERE regexp_replace(phone,'[^0-9]','','g') LIKE %s AND is_duplicate=0 LIMIT 1""",
            (f"%{phone}",))
        r = cur.fetchone()
        if r: return True, r[0], "phone"
    if email:
        cur.execute("SELECT id FROM candidates WHERE lower(trim(email_addr))=%s AND is_duplicate=0 LIMIT 1", (email,))
        r = cur.fetchone()
        if r: return True, r[0], "email"
    return False, None, None

def check_missing(row):
    m = []
    if not row.get("candidate_name"): m.append("name")
    if not row.get("phone") and not row.get("email_addr"): m.append("phone/email")
    elif not row.get("phone"):       m.append("phone")
    elif not row.get("email_addr"):  m.append("email")
    return m

# ── Upload handler ────────────────────────────────────────────────────────────
def process_upload(recruiter_name, recruiter_email, role_override, excel_file, cv_file_list):
    batch_id = hashlib.md5(f"{recruiter_email}{datetime.now().isoformat()}".encode()).hexdigest()[:10]
    result   = {"added":0,"duplicates":0,"missing":0,"errors":[],"cv_warnings":[]}

    fname = excel_file.filename.lower()
    data  = excel_file.read()
    if   fname.endswith(".csv"):           rows, err = parse_csv(data, role_override)
    elif fname.endswith((".xlsx",".xls")): rows, err = parse_xlsx(data, role_override)
    else: return {"error":"Please upload a .xlsx or .csv file"}
    if err:   return {"error": err}
    if not rows: return {"error":"No rows found. Check the file has a header row with data below it."}

    # Upload CVs
    saved_cvs = []
    for cv in cv_file_list:
        if cv and cv.filename:
            orig, url, pub_id = upload_cv(cv, batch_id)
            if orig: saved_cvs.append((orig, url, pub_id))

    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO team_members (name,email,is_fixed) VALUES (%s,%s,0) ON CONFLICT (email) DO NOTHING",
                (recruiter_name, recruiter_email.lower()))

    matched_ids = set()
    for row in rows:
        if not any(str(v).strip() for v in row.values()): continue
        cname = row.get("candidate_name","").strip()
        cv_orig, cv_url, cv_pub, cv_ok = match_cv(cname, saved_cvs)
        if cv_pub: matched_ids.add(cv_pub)
        is_dup, dup_id, dup_why = check_dup(conn, row)
        missing = check_missing(row)

        cur.execute("""INSERT INTO candidates
            (upload_batch,recruiter_name,recruiter_email,role_name,candidate_name,
             email_addr,phone,current_company,current_role,experience_years,key_skills,
             notice_period,current_salary,expected_salary,current_location,
             preferred_location,remarks,cv_filename,cv_url,cv_public_id,
             is_duplicate,duplicate_of,missing_info)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id""",
            (batch_id,recruiter_name,recruiter_email.lower(),
             row.get("role_name",""),cname,row.get("email_addr",""),row.get("phone",""),
             row.get("current_company",""),row.get("current_role",""),
             row.get("experience_years",""),row.get("key_skills",""),
             row.get("notice_period",""),row.get("current_salary",""),
             row.get("expected_salary",""),row.get("current_location",""),
             row.get("preferred_location",""),row.get("remarks",""),
             cv_orig,cv_url,cv_pub,1 if is_dup else 0,dup_id,
             ",".join(missing) if missing else None))
        cid = cur.fetchone()[0]

        if is_dup:
            cur.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (%s,%s,%s,%s)",
                ("duplicate",f"Duplicate: {cname or '?'} matched via {dup_why} (original ID #{dup_id})",cid,recruiter_email))
            result["duplicates"] += 1
        if missing:
            cur.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (%s,%s,%s,%s)",
                ("missing_info",f"Missing {', '.join(missing)} for {cname or 'unnamed'}",cid,recruiter_email))
            result["missing"] += 1
        result["added"] += 1

    for orig, url, pid in unmatched_cvs(saved_cvs, matched_ids):
        cur.execute("INSERT INTO alerts (alert_type,message,recruiter_email) VALUES (%s,%s,%s)",
            ("cv_mismatch",
             f"CV '{orig}' by {recruiter_name} doesn't match any candidate name in the Excel. Please re-upload with correct filename.",
             recruiter_email))
        result["cv_warnings"].append(orig)

    cur.execute("""INSERT INTO upload_log
        (batch_id,recruiter_name,recruiter_email,filename,candidates_added,duplicates_found,missing_count)
        VALUES (%s,%s,%s,%s,%s,%s,%s)""",
        (batch_id,recruiter_name,recruiter_email.lower(),excel_file.filename,
         result["added"],result["duplicates"],result["missing"]))

    today = date.today().isoformat()
    cur.execute("SELECT DISTINCT recruiter_email FROM candidates WHERE created_at::date=%s",(today,))
    submitted = {r[0] for r in cur.fetchall()}
    cur.execute("SELECT name,email FROM team_members")
    for n,e in cur.fetchall():
        if e not in submitted:
            cur.execute("SELECT 1 FROM alerts WHERE alert_type='no_submission' AND recruiter_email=%s AND created_at::date=%s",(e,today))
            if not cur.fetchone():
                cur.execute("INSERT INTO alerts (alert_type,message,recruiter_email) VALUES (%s,%s,%s)",
                    ("no_submission",f"{n} has not submitted any profiles today",e))

    conn.commit(); cur.close(); conn.close()
    return result

# ── Template ──────────────────────────────────────────────────────────────────
def make_template():
    if not XLSX_OK: return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Candidates"
    headers = ["Candidate Name","Email","Phone","Current Company","Current Role",
               "Experience (Years)","Key Skills","Notice Period","Current Salary",
               "Expected Salary","Current Location","Preferred Location","Role","Remarks"]
    hfill = PatternFill("solid",fgColor="1C2030"); hfont = Font(bold=True,color="E8643A")
    for i,h in enumerate(headers,1):
        c = ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
        c.alignment=Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(len(h)+4,18)
    hints=["← Full name","Email address","10-digit mobile","Current employer","Job title",
           "e.g. 5 years","Comma separated","e.g. 30 days","e.g. 12 LPA","e.g. 18 LPA",
           "Current city","Preferred city","Role being screened","Any notes"]
    ifill=PatternFill("solid",fgColor="0D1017"); ifont=Font(italic=True,color="6B7494")
    for i,h in enumerate(hints,1):
        c=ws.cell(row=2,column=i,value=h); c.font=ifont; c.fill=ifill
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── Export config ─────────────────────────────────────────────────────────────
EXPORT_COLS = ["candidate_name","email_addr","phone","current_company","current_role",
               "experience_years","key_skills","notice_period","current_salary",
               "expected_salary","current_location","preferred_location",
               "role_name","recruiter_name","remarks","created_at"]
EXPORT_HDR  = ["Candidate Name","Email","Phone","Current Company","Current Role",
               "Experience","Key Skills","Notice Period","Current Salary","Expected Salary",
               "Current Location","Preferred Location","Role","Recruiter","Remarks","Date Added"]

def build_query(args):
    role=args.get("role",""); sender=args.get("sender",""); loc=args.get("location","")
    notice=args.get("notice",""); q=args.get("q",""); show_d=args.get("show_dups","0")
    sql="SELECT * FROM candidates WHERE 1=1"; p=[]
    if role:   sql+=" AND role_name=%s";   p.append(role)
    if sender: sql+=" AND recruiter_email=%s"; p.append(sender)
    if loc:    sql+=" AND (current_location ILIKE %s OR preferred_location ILIKE %s)"; p+=[f"%{loc}%"]*2
    if notice: sql+=" AND notice_period ILIKE %s"; p.append(f"%{notice}%")
    if show_d=="0": sql+=" AND is_duplicate=0"
    if q:
        sql+=" AND (candidate_name ILIKE %s OR key_skills ILIKE %s OR current_company ILIKE %s OR phone ILIKE %s OR email_addr ILIKE %s)"
        p+=[f"%{q}%"]*5
    sql+=" ORDER BY created_at DESC"; return sql, p

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index(): return render_template("index.html")

@app.route("/upload")
def upload_page(): return render_template("upload.html")

@app.route("/healthz")
def health(): return "ok"

@app.route("/api/upload", methods=["POST"])
def api_upload():
    name=request.form.get("recruiter_name","").strip()
    email=request.form.get("recruiter_email","").strip().lower()
    role=request.form.get("role_override","").strip()
    if not name or not email: return jsonify({"error":"Name and email required"}),400
    if "excel_file" not in request.files: return jsonify({"error":"No file attached"}),400
    return jsonify(process_upload(name,email,role,request.files["excel_file"],request.files.getlist("cv_files")))

@app.route("/api/template")
def api_template():
    buf=make_template()
    if not buf: return "openpyxl not installed",500
    return send_file(buf,download_name="candidate_template.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/candidates")
def api_candidates():
    conn=get_db(); cur=conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    sql,p=build_query(request.args); cur.execute(sql,p)
    rows=cur.fetchall(); cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/candidates/export")
def export_candidates():
    fmt=request.args.get("fmt","csv")
    conn=get_db(); cur=conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    sql,p=build_query(request.args); cur.execute(sql,p)
    rows=[dict(r) for r in cur.fetchall()]; cur.close(); conn.close()

    if fmt=="xlsx" and XLSX_OK:
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Candidates"
        hfill=PatternFill("solid",fgColor="1C2030"); hfont=Font(bold=True,color="E8643A")
        for i,h in enumerate(EXPORT_HDR,1):
            c=ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
            ws.column_dimensions[c.column_letter].width=max(len(h)+4,16)
        for ri,row in enumerate(rows,2):
            for ci,col in enumerate(EXPORT_COLS,1):
                v=row.get(col,""); ws.cell(row=ri,column=ci,value=str(v) if v else "")
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,download_name=f"candidates_{date.today()}.xlsx",as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    buf=io.StringIO(); writer=csv.writer(buf); writer.writerow(EXPORT_HDR)
    for row in rows: writer.writerow([str(row.get(c,"") or "") for c in EXPORT_COLS])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),download_name=f"candidates_{date.today()}.csv",
                     as_attachment=True,mimetype="text/csv")

@app.route("/api/filters")
def api_filters():
    conn=get_db(); cur=conn.cursor()
    cur.execute("SELECT DISTINCT role_name FROM candidates WHERE role_name!='' ORDER BY role_name")
    roles=[r[0] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT recruiter_email,recruiter_name FROM candidates WHERE recruiter_email!='' ORDER BY recruiter_name")
    senders=[{"recruiter_email":r[0],"recruiter_name":r[1]} for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT current_location FROM candidates WHERE current_location!='' ORDER BY current_location")
    locs=[r[0] for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify({"roles":roles,"senders":senders,"locations":locs})

@app.route("/api/stats")
def api_stats():
    conn=get_db(); cur=conn.cursor()
    cur.execute("SELECT COUNT(*) FROM candidates WHERE is_duplicate=0"); total=cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM candidates WHERE is_duplicate=1"); dups=cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT role_name) FROM candidates WHERE is_duplicate=0"); roles=cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM alerts WHERE is_read=0"); unread=cur.fetchone()[0]
    cur.execute("SELECT uploaded_at FROM upload_log ORDER BY id DESC LIMIT 1"); last=cur.fetchone()
    cur.close(); conn.close()
    return jsonify({"total":total,"duplicates":dups,"roles":roles,"unread_alerts":unread,
                    "last_upload":str(last[0]) if last else None})

@app.route("/api/alerts")
def api_alerts():
    conn=get_db(); cur=conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    atype=request.args.get("type","")
    sql="SELECT * FROM alerts WHERE 1=1"; p=[]
    if atype: sql+=" AND alert_type=%s"; p.append(atype)
    sql+=" ORDER BY created_at DESC LIMIT 150"
    cur.execute(sql,p); rows=cur.fetchall(); cur.close(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/alerts/read",methods=["POST"])
def mark_all_read():
    conn=get_db(); cur=conn.cursor(); cur.execute("UPDATE alerts SET is_read=1")
    conn.commit(); cur.close(); conn.close(); return jsonify({"ok":True})

@app.route("/api/alerts/<int:aid>/read",methods=["POST"])
def mark_one_read(aid):
    conn=get_db(); cur=conn.cursor()
    cur.execute("UPDATE alerts SET is_read=1 WHERE id=%s",(aid,))
    conn.commit(); cur.close(); conn.close(); return jsonify({"ok":True})

@app.route("/api/reporting/recruiter_performance")
def recruiter_performance():
    conn=get_db(); cur=conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    days=int(request.args.get("days",30)); since=(date.today()-timedelta(days=days)).isoformat()
    cur.execute("""SELECT recruiter_name,recruiter_email,COUNT(*) as total,
        SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_candidates,
        SUM(CASE WHEN is_duplicate=1 THEN 1 ELSE 0 END) as duplicates,
        COUNT(DISTINCT role_name) as roles_worked,
        COUNT(DISTINCT created_at::date) as active_days
        FROM candidates WHERE created_at>=%s
        GROUP BY recruiter_email,recruiter_name ORDER BY unique_candidates DESC""",(since,))
    rows=cur.fetchall(); cur.close(); conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/reporting/daily_trend")
def daily_trend():
    conn=get_db(); cur=conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    days=int(request.args.get("days",14)); since=(date.today()-timedelta(days=days)).isoformat()
    cur.execute("""SELECT created_at::date as day,COUNT(*) as total,
        SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_c,
        COUNT(DISTINCT recruiter_email) as active_recruiters
        FROM candidates WHERE created_at>=%s GROUP BY created_at::date ORDER BY day""",(since,))
    rows=cur.fetchall(); cur.close(); conn.close()
    return jsonify([{"day":str(r["day"]),"total":r["total"],"unique_c":r["unique_c"],
                     "active_recruiters":r["active_recruiters"]} for r in rows])

@app.route("/api/reporting/no_submission_today")
def no_submission_today():
    conn=get_db(); cur=conn.cursor(); today=date.today().isoformat()
    cur.execute("SELECT DISTINCT recruiter_email FROM candidates WHERE created_at::date=%s",(today,))
    submitted={r[0] for r in cur.fetchall()}
    cur.execute("SELECT name,email FROM team_members")
    missing=[{"name":r[0],"email":r[1]} for r in cur.fetchall() if r[1] not in submitted]
    cur.close(); conn.close(); return jsonify(missing)

@app.route("/api/team",methods=["GET"])
def get_team():
    conn=get_db(); cur=conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM team_members ORDER BY name")
    rows=cur.fetchall(); cur.close(); conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/team",methods=["POST"])
def add_team():
    d=request.json; conn=get_db(); cur=conn.cursor()
    cur.execute("INSERT INTO team_members (name,email,is_fixed) VALUES (%s,%s,1) ON CONFLICT (email) DO NOTHING",
                (d["name"],d["email"].lower()))
    conn.commit(); cur.close(); conn.close(); return jsonify({"ok":True})

@app.route("/api/team/<int:tid>",methods=["DELETE"])
def del_team(tid):
    conn=get_db(); cur=conn.cursor()
    cur.execute("DELETE FROM team_members WHERE id=%s",(tid,))
    conn.commit(); cur.close(); conn.close(); return jsonify({"ok":True})

@app.route("/api/candidate/<int:cid>",methods=["PATCH"])
def update_candidate(cid):
    data=request.json
    allowed=["candidate_name","email_addr","phone","current_company","current_role",
             "experience_years","key_skills","notice_period","current_salary",
             "expected_salary","current_location","preferred_location","remarks","role_name"]
    sets=", ".join(f"{k}=%s" for k in data if k in allowed)
    vals=[v for k,v in data.items() if k in allowed]+[cid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db(); cur=conn.cursor()
    cur.execute(f"UPDATE candidates SET {sets} WHERE id=%s",vals)
    conn.commit(); cur.close(); conn.close(); return jsonify({"ok":True})

@app.route("/api/candidate/<int:cid>",methods=["DELETE"])
def del_candidate(cid):
    conn=get_db(); cur=conn.cursor()
    cur.execute("DELETE FROM candidates WHERE id=%s",(cid,))
    conn.commit(); cur.close(); conn.close(); return jsonify({"ok":True})

if __name__=="__main__":
    init_db()
    print("✅  HR Guru ATS → http://localhost:5000")
    app.run(debug=False,host="0.0.0.0",port=int(os.getenv("PORT",5000)))
